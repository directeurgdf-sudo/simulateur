
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import json

st.set_page_config(page_title="Simulateur Excel en ligne", page_icon="🧮")

st.title("🧮 Simulateur Excel en ligne")
st.caption("Charge un classeur Excel et expose des champs simples via une interface web.")

st.markdown(
    """
**Comment ça marche ?**
1) Uploade ton fichier Excel (XLSX).  
2) Le fichier `config.json` (dans le repo) décrit quels **champs d'entrée** écrire dans le fichier et quelles **cellules de sortie** lire.  
3) Clique **Calculer & Générer** pour écrire les valeurs d'entrée dans le classeur et lire les résultats (valeurs en cache des formules).  
4) Télécharge la version mise à jour de l'Excel.

> ⚠️ Les moteurs Excel natifs ne tournent pas côté serveur. Si tes formules ne sont pas recalculées par Excel lui‑même,
l'application lira la **valeur en cache** stockée dans le fichier. Pour forcer une mise à jour des formules, ouvre puis sauvegarde le fichier dans Excel/LibreOffice avant de le réimporter ici.
"""
)

# Load config
default_config = {
    "inputs": [
        # Exemple: {"label": "Chiffre d'affaires", "cell": "B3", "type": "number", "min": 0.0, "max": 1_000_000.0, "step": 1000.0, "sheet": "Feuil1"}
    ],
    "outputs": [
        # Exemple: {"label": "Résultat net", "cell": "E10", "sheet": "Feuil1"}
    ]
}

try:
    with open("config.json", "r", encoding="utf-8") as f:
        cfg = json.load(f)
except Exception:
    cfg = default_config

uploaded = st.file_uploader("📤 Uploader un fichier Excel (.xlsx)", type=["xlsx"])

# Sidebar: inputs
st.sidebar.header("Paramètres")
inputs_state = {}
if cfg.get("inputs"):
    st.sidebar.subheader("Entrées")
    for field in cfg["inputs"]:
        label = field.get("label", field.get("cell", "Entrée"))
        sheet = field.get("sheet")
        key = f"{sheet or ''}:{field.get('cell','')}:{label}"
        t = field.get("type", "number")
        if t == "number":
            v = st.sidebar.number_input(
                label,
                min_value=field.get("min", 0.0),
                max_value=field.get("max", 1_000_000.0),
                value=field.get("default", 0.0),
                step=field.get("step", 1.0),
                key=key,
            )
        elif t == "int":
            v = st.sidebar.number_input(
                label,
                min_value=int(field.get("min", 0)),
                max_value=int(field.get("max", 1_000_000)),
                value=int(field.get("default", 0)),
                step=int(field.get("step", 1)),
                key=key,
            )
        elif t == "text":
            v = st.sidebar.text_input(label, value=str(field.get("default","")), key=key)
        elif t == "select":
            options = field.get("options", [])
            v = st.sidebar.selectbox(label, options=options, index=0 if options else None, key=key)
        else:
            v = st.sidebar.text_input(label, value=str(field.get("default","")), key=key)
        inputs_state[key] = {"value": v, "cell": field.get("cell"), "sheet": sheet}

calculate = st.button("🚀 Calculer & Générer", disabled=(uploaded is None))

if uploaded is not None:
    st.success("Fichier chargé.")
    # Preview sheets
    try:
        xl = pd.ExcelFile(uploaded)
        st.write("**Aperçu des feuilles :**", xl.sheet_names)
        # show preview of first sheet
        try:
            df = xl.parse(xl.sheet_names[0]).head(10)
            st.dataframe(df)
        except Exception:
            pass
    except Exception as e:
        st.error(f"Impossible de lire le classeur: {e}")

    if calculate:
        uploaded.seek(0)
        wb = load_workbook(uploaded, data_only=False)  # keep formulas
        # Write inputs
        for _, f in inputs_state.items():
            cell = f["cell"]
            sheet = f.get("sheet")
            value = f["value"]
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            ws[cell] = value

        # Save to buffer
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        # Re-open with data_only=True to read cached formula results
        wb2 = load_workbook(bio, data_only=True)
        results = []
        for field in cfg.get("outputs", []):
            label = field.get("label", field.get("cell", "Sortie"))
            sheet = field.get("sheet")
            cell = field.get("cell")
            ws = wb2[sheet] if sheet and sheet in wb2.sheetnames else wb2.active
            val = ws[cell].value
            results.append((label, val))

        if results:
            st.subheader("📊 Résultats (valeurs lues)")
            for label, val in results:
                st.metric(label, val)

        st.download_button(
            "📥 Télécharger l'Excel mis à jour",
            data=bio.getvalue(),
            file_name="simulateur_mis_a_jour.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

st.divider()
with st.expander("⚙️ Configurer les champs (config.json)"):
    st.markdown(
        """
Crée/édite un fichier **config.json** à la racine du repo :
```
{
  "inputs": [
    {"label": "Chiffre d'affaires", "cell": "B3", "type": "number", "min": 0, "max": 1000000, "step": 1000, "default": 50000, "sheet": "Feuil1"},
    {"label": "Taux TVA", "cell": "B4", "type": "number", "min": 0, "max": 1, "step": 0.01, "default": 0.2, "sheet": "Feuil1"},
    {"label": "Type d'activité", "cell": "B5", "type": "select", "options": ["Gîte","CH"], "default": "Gîte", "sheet": "Feuil1"}
  ],
  "outputs": [
    {"label": "Résultat net", "cell": "E10", "sheet": "Feuil1"},
    {"label": "Marge", "cell": "E11", "sheet": "Feuil1"}
  ]
}
```
- **inputs** : champs affichés dans la barre latérale, écrits dans les cellules correspondantes.
- **outputs** : cellules lues après écriture (valeurs en cache si formules).

Astuce : si tes formules ne se recalculent pas côté serveur, ouvre le fichier dans Excel/LibreOffice, fais *Recalculer*, puis sauvegarde avant de le réutiliser ici.
"""
    )
